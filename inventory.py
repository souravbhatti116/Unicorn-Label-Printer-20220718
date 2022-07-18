import ctypes
import os
import shutil
import tkinter as tk
from datetime import datetime
from tkinter import *
from tkinter import messagebox, simpledialog, ttk, Button, filedialog
from tkinter.ttk import Combobox, Treeview

import pandas as pd
import openpyxl


dic_names = ["Black Plug Gateway, Ext Ant, CAN, Type A",
             "Black Plug Gateway, Ext Ant, EU, Type C",
             "Black Plug Gateway, Ext Ant, EU, Type G",
             "Black Plug Gateway, Ext Ant, US, ATT",
             "Black Plug Gateway, Ext Ant, US, TMO",
             "Black Plug Gateway, Ext Ant, US, VZN",
             "Green Wallplug US",
             "Green Wallplug, CAN, Type B",
             "Green Wallplug, EU, Type E",
             "Green Wallplug, EU, Type G"]

dic_numbers = ["GBP-2002-CAN-EXA",
               "GBP-2002-EUC-EXA",
               "GBP-2002-EUG-EXA",
               "GBP-2002-0A2-ATT",
               "GBP-2002-0A2-TMO",
               "GBP-2002-0A2-VZN",
               "PGW-2003-0A1",
               "PGW-2003-CAN",
               "PGW-2003-EUE",
               "PGW-2003-EUG"]


# dic_data = dict(zip(dic_names, dic_num


def inventoryCheckOut():
    root = tk.Tk()
    root.geometry("870x450+300+200")
    root.configure(background="light blue")
    root.title("Check Out Inventory")
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
    #messagebox.showinfo("Attention", "This Window is for Inventory Check-out.")

    style = ttk.Style()
    style.theme_use('clam')

    # Label Creation
    namelabel = tk.Label(root, text="Product Name :", font=('Times new roman', 13, 'bold'), bg='Light Blue')
    qrcodelabel = tk.Label(root, text="QR Code :", font=('Times new roman', 13, 'bold'), bg='Light Blue')

    # Labels placement
    namelabel.place(x=230, y=25)
    qrcodelabel.place(x=230, y=90)

    # StringVar() to store the values from User input.
    name = StringVar()
    number = StringVar()
    qrcode = StringVar()
    status = StringVar()

    # Combobox creation
    nameComboBox = Combobox(root, state='readonly', textvariable=name, width=35, values=dic_names,
                            postcommand=lambda: selnumber())  # Put the values
    # numComboBox = Combobox(root, textvariable=number, width=35, values=dic_numbers)  # Put the values
    qrEntryBox = tk.Entry(root, textvariable=qrcode, width=38)

    # Label Placement
    nameComboBox.place(x=400, y=25)
    qrEntryBox.place(x=400, y=90)

    # Tree view
    displaydata = Treeview(root, columns=['Product Name', 'Product Number', 'QR Code', 'Status'], show='headings',
                           height=10)
    displaydata.place(x=30, y=170)
    displaydata.column("# 1", anchor=CENTER)
    displaydata.heading("# 1", text="Product Name")
    displaydata.column("# 2", anchor=CENTER)
    displaydata.heading("# 2", text="Product Number")
    displaydata.column("# 3", anchor=CENTER)
    displaydata.heading("# 3", text="QR Code")
    displaydata.column("# 4", anchor=CENTER)
    displaydata.heading("# 4", text="Status")

    def selnumber():
        if name.get() == "Black Plug Gateway, Ext Ant, CAN, Type A":
            number.set("GBP-2002-CAN-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type C":
            number.set("GBP-2002-EUC-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type G":
            number.set("GBP-2002-EUG-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, US, ATT":
            number.set("GBP-2002-0A2-ATT")
        if name.get() == "Black Plug Gateway, Ext Ant, US, TMO":
            number.set("GBP-2002-0A2-TMO")
        if name.get() == "Black Plug Gateway, Ext Ant, US, VZN":
            number.set("GBP-2002-0A2-VZN")
        if name.get() == "Green Wallplug US":
            number.set("PGW-2003-0A1")
        if name.get() == "Green Wallplug, CAN, Type B":
            number.set("PGW-2003-CAN")
        if name.get() == "Green Wallplug, EU, Type E":
            number.set("PGW-2003-EUE")
        if name.get() == "Green Wallplug, EU, Type G":
            number.set("PGW-2003-EUG")

    def cleartext():
        qrEntryBox.delete(0, END)

    def datacheck():
        if qrcode.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No QR code entered')
            root.destroy()
            inventoryCheckOut()
        elif name.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No Product Name entered')
            root.destroy()
            inventoryCheckOut()

    def save_to_excel_out():
        tstamp = str(datetime.now().strftime("%Y-%m-%d %I:%M%p"))
        f = openpyxl.load_workbook('InventoryCheckOut.xlsx')
        sheet = f.active
        maxrow = sheet.max_row

        data = ({"Product Name": [name.get()], "Product Number": [number.get()], "QR Code": [qrcode.get()],
                 "TimeStamp": [tstamp], "Status": [status.get()]})

        dataframe = pd.DataFrame(data)

        with pd.ExcelWriter('InventoryCheckOut.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            dataframe.to_excel(writer, startrow=maxrow, startcol=0, header=None, index=False)

    root.bind('<Return>', lambda event: checkout())

    def checkout():
        datacheck()
        duplicatesOut()
        status.set("Checked Out")
        selnumber()
        displaydata.insert('', END,
                           values=[name.get(), number.get(), qrcode.get(), status.get()])
        save_to_excel_out()  # Just for saving into a file SavedOnly xlsx.
        cleartext()

    # checkoutbtn = Button(root, text="Check Out", command=lambda: checkout())
    # checkoutbtn.place(x=400, y=125)

    def popupwindow():
        pilotname = simpledialog.askstring("Enter Name of the Product.", "Enter Name of the Product.", parent=root)
        name.set(pilotname)
        pilotnum = simpledialog.askstring("Enter number of the Product.", "Enter Number of the Product.", parent=root)
        number.set(pilotnum)

    manualbtn = Button(root, text="Manual Entry", command=lambda: popupwindow(), font=('Times new roman', 15),
                       relief='ridge')
    manualbtn.place(x=730, y=50)

    def duplicatesOut():
        currentQrCode = qrcode.get()
        # currentName = name.get()
        # currentNumber = productNumber.get()
        wb = openpyxl.load_workbook('inventoryCheckOut.xlsx', read_only=True)
        cs = wb.active

        for row in cs.rows:
            for cell in row:
                if currentQrCode == cell.value:
                    messagebox.showerror("Error", "Duplicate Entry!")
                    root.destroy()
                    inventoryCheckOut()

        wb.close()

    def switchwindow():
        root.destroy()
        import main
        main.main()

    photo = PhotoImage(file='Home_icon_black.png')

    switchbtn = Button(root, image=photo, height=40, width=40, text="Home", command=lambda: switchwindow())

    switchbtn.place(x=70, y=20)

    def displaypreviousdata():

        file = filedialog.askopenfilename(title="Open a File", filetype=(("xlsx files", ".*xlsx"), ("All Files", "*.")))

        df = pd.read_excel(file)

        for item in displaydata.get_children():
            displaydata.delete(item)

        for cell in range(len(df)):
            displaydata.insert("", END, values=list(df.loc[cell]))

    historybtn = Button(root, text="History", command=lambda: displaypreviousdata(), font=('Times new roman', 10),
                        relief='ridge')
    historybtn.place(x=70, y=90)

    def cleartree():
         for item in displaydata.get_children():
          displaydata.delete(item)

    cleartreebtn = Button(root, text="Clear Data", command= lambda : cleartree())
    cleartreebtn.place(x= 400, y= 410)

    def savefile():

        source = "InventoryCheckOut.xlsx"
        sourcefile = os.path.abspath(source)

        destination = filedialog.askopenfilename()
        destinationfile = os.path.abspath(destination)

        shutil.copy(sourcefile, destinationfile)

        # source = openpyxl.load_workbook("LabelPrintLog.xlsx")
        # sheet = source.active


    savebtn = Button(root, text="End Of The Day", command=lambda: savefile(), font=('Times new roman', 10),
                        relief='ridge')
    savebtn.place(x=70, y=130)


    root.mainloop()


# ###################################################################################################################

def inventoryCheckIN():
    root = tk.Tk()
    root.geometry("870x450+300+200")
    root.configure(background="orange")
    root.title("Check In Inventory")
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
    #messagebox.showinfo("Attention", "This Window is for Inventory Check in.")

    style = ttk.Style()
    style.theme_use('clam')

    # Label Creation
    namelabel = tk.Label(root, text="Product Name :", font=('Times new roman', 13, 'bold'), bg='orange')
    qrcodelabel = tk.Label(root, text="QR Code:", font=('Times new roman', 13, 'bold'), bg='orange')

    # Labels placement
    namelabel.place(x=230, y=25)
    qrcodelabel.place(x=230, y=90)

    # StringVar() to store the values from User input.
    name = StringVar()
    number = StringVar()
    qrcode = StringVar()
    status = StringVar()

    # Combobox creation
    nameComboBox = Combobox(root, state='readonly', textvariable=name, width=35, values=dic_names,
                            postcommand=lambda: selnumber())  # Put the values
    # numComboBox = Combobox(root, textvariable=number, width=35, values=dic_numbers)  # Put the values
    qrEntryBox = tk.Entry(root, textvariable=qrcode, width=38)

    # Label Placement
    nameComboBox.place(x=400, y=25)
    qrEntryBox.place(x=400, y=90)

    # Tree view
    displaydata = Treeview(root, columns=['Product Name', 'Product Number', 'QR Code', 'Status'], show='headings',
                           height=10)
    displaydata.place(x=30, y=170)
    displaydata.column("# 1", anchor=CENTER)
    displaydata.heading("# 1", text="Product Name")
    displaydata.column("# 2", anchor=CENTER)
    displaydata.heading("# 2", text="Product Number")
    displaydata.column("# 3", anchor=CENTER)
    displaydata.heading("# 3", text="QR Code")
    displaydata.column("# 4", anchor=CENTER)
    displaydata.heading("# 4", text="Status")

    def selnumber():
        if name.get() == "Black Plug Gateway, Ext Ant, CAN, Type A":
            number.set("GBP-2002-CAN-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type C":
            number.set("GBP-2002-EUC-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type G":
            number.set("GBP-2002-EUG-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, US, ATT":
            number.set("GBP-2002-0A2-ATT")
        if name.get() == "Black Plug Gateway, Ext Ant, US, TMO":
            number.set("GBP-2002-0A2-TMO")
        if name.get() == "Black Plug Gateway, Ext Ant, US, VZN":
            number.set("GBP-2002-0A2-VZN")
        if name.get() == "Green Wallplug US":
            number.set("PGW-2003-0A1")
        if name.get() == "Green Wallplug, CAN, Type B":
            number.set("PGW-2003-CAN")
        if name.get() == "Green Wallplug, EU, Type E":
            number.set("PGW-2003-EUE")
        if name.get() == "Green Wallplug, EU, Type G":
            number.set("PGW-2003-EUG")

    def cleartext():
        qrEntryBox.delete(0, END)

    def datacheck():
        if qrcode.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No QR code entered')
            root.destroy()
            inventoryCheckIN()
        elif name.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No Product Name entered')
            root.destroy()
            inventoryCheckIN()

    def save_to_excel_in():
        tstamp = str(datetime.now().strftime("%Y-%m-%d %I:%M%p"))
        f = openpyxl.load_workbook('InventoryCheckIn.xlsx')
        sheet = f.active
        maxrow = sheet.max_row

        data = ({"Product Name": [name.get()], "Product Number": [number.get()], "QR Code": [qrcode.get()],
                 "TimeStamp": [tstamp], "Status": [status.get()]})

        dataframe = pd.DataFrame(data)

        with pd.ExcelWriter('InventoryCheckIn.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            dataframe.to_excel(writer, startrow=maxrow, startcol=0, header=None, index=False)

    def checkin():
        datacheck()
        DuplicatesIn()
        status.set("Checked in")
        selnumber()
        displaydata.insert('', END,
                           values=[name.get(), number.get(), qrcode.get(), status.get()])
        save_to_excel_in()  # Just for saving into a file SavedOnly xlsx.
        cleartext()

    # checkinbtn = Button(root, text="Check In", command=lambda: checkin())
    # checkinbtn.place(x=570, y=125)
    root.bind('<Return>', lambda event: checkin())

    def popupwindow():
        pilotname = simpledialog.askstring("Enter Name of the Product.", "Enter Name of the Product.", parent=root)
        name.set(pilotname)
        pilotnum = simpledialog.askstring("Enter number of the Product.", "Enter Number of the Product.", parent=root)
        number.set(pilotnum)

    manualbtn = Button(root, text="Manual Entry", command=lambda: popupwindow(), font=('Times new roman', 15),
                       relief='ridge')
    manualbtn.place(x=730, y=50)

    def DuplicatesIn():
        currentQrCode = qrcode.get()
        # currentName = name.get()
        # currentNumber = productNumber.get()
        wb = openpyxl.load_workbook('inventoryCheckIn.xlsx', read_only=True)
        cs = wb.active

        for row in cs.rows:
            for cell in row:
                if currentQrCode == cell.value:
                    messagebox.showerror("Error", "Duplicate Entry!")
                    root.destroy()
                    inventoryCheckIN()

        wb.close()

    def switchwindow():
        root.destroy()
        import main
        main.main()

    photo = PhotoImage(file='Home_icon_black.png')

    switchbtn: Button = Button(root, image=photo, height= 40, width=40, text="Home", command=lambda: switchwindow())

    switchbtn.place(x=70, y=20)

    def displaypreviousdata():

        file = filedialog.askopenfilename(title="Open a File", filetype=(("xlsx files", ".*xlsx"), ("All Files", "*.")))

        df = pd.read_excel(file)

        for item in displaydata.get_children():
            displaydata.delete(item)

        for cell in range(len(df)):
            displaydata.insert("", END, values=list(df.loc[cell]))

    historybtn = Button(root, text="History", command=lambda: displaypreviousdata(), font=('Times new roman', 10),
                        relief='ridge')
    historybtn.place(x=70, y=90)

    def cleartree():
         for item in displaydata.get_children():
          displaydata.delete(item)

    cleartreebtn = Button(root, text="Clear Data", command= lambda : cleartree())
    cleartreebtn.place(x= 400, y= 410)

    def savefile():

        source = "InventoryCheckIn.xlsx"
        sourcefile = os.path.abspath(source)

        destination = filedialog.askopenfilename()
        destinationfile = os.path.abspath(destination)

        shutil.copy(sourcefile, destinationfile)

        # source = openpyxl.load_workbook("LabelPrintLog.xlsx")
        # sheet = source.active

    savebtn = Button(root, text="End Of the Day", command=lambda: savefile(), font=('Times new roman', 10),
                     relief='ridge')
    savebtn.place(x=70, y=130)

    root.mainloop()
