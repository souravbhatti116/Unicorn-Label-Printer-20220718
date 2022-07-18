import ctypes
from genericpath import exists
import json
import os
import shutil
import sys
from tkinter import *
import tkinter as tk
from tkinter import ttk, filedialog, Label
from tkinter import messagebox
from tkinter.ttk import Combobox, Treeview
from turtle import update, width
from numpy import pad

from openpyxl.chart import label
from unicodedata import name
import openpyxl
import requests
# import pywin
from win32com.client import Dispatch
import pathlib
import pandas as pd
from datetime import datetime
from tkinter import simpledialog

dic_names = ["Black Plug Gateway, Ext Ant, CAN, Type A",
             "Black Plug Gateway, Ext Ant, EU, Type C",
             "Black Plug Gateway, Ext Ant, EU, Type G",
             "Black Plug Gateway, Ext Ant, US, ATT",
             "Black Plug Gateway, Ext Ant, US, TMO",
             "Black Plug Gateway, Ext Ant, US, VZN",
             "Green Wallplug US",
             "Green Wallplug, CAN, Type B",
             "Green Wallplug, EU, Type E",
             "Green Wallplug, EU, Type G",
             "Black Plug Gateway, Ext Ant, ASIA, Type A",
            "Black Plug Gateway, Ext Ant, IND, Type C",
            "Black Plug Gateway, Ext Ant, HK, Type G",
            "Black Plug Gateway, Ext Ant, NZ, Type I",
            "Black Plug Gateway, Ext Ant, AUS, Type I",
            "Black Plug Gateway, Ext Ant, SG, Type C",
            "Black Plug Gateway, Ext Ant, KOR, Type C",
            "Black Plug Gateway, Ext Ant, MEX, Type C",
            "Green Wallplug, ASIA, Type B",
            "Green Wallplug, IND, Type E",
            "Green Wallplug, HK, Type G",
            "Green Wallplug NZ, Type I",
            "Green Wallplug, AUS, Type I",
            "Green Wallplug, SG, Type E",
            "Green Wallplug, KOR, Type E",
            "Green Wallplug, MEX, Type E",]

# dic_numbers = ["GBP-2002-CAN-EXA",
#                "GBP-2002-EUC-EXA",
#                "GBP-2002-EUG-EXA",
#                "GBP-2002-0A2-ATT",
#                "GBP-2002-0A2-TMO",
#                "GBP-2002-0A2-VZN",
#                "PGW-2003-0A1",
#                "PGW-2003-CAN",
#                "PGW-2003-EUE",
#                "PGW-2003-EUG"]


# dic_data = dict(zip(dic_names, dic_numbers))


def printing():
    
    root = tk.Tk()
    root.geometry("870x700+300+200")
    root.configure(background="aquamarine")
    root.title("Print Labels")
    root.resizable(0,0  )
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
   # messagebox.showinfo("Attention", "This Window is for printing out labels only.")

    style = ttk.Style()
    style.theme_use('clam')

    menuBar = Menu(root, font=("Times new roman", 13, 'bold'))
    menuOptions = Menu(menuBar,font=("Times new roman", 13, 'bold'), tearoff=0)
    
    menuOptions.add_command(label="Manual Entry", command= lambda:popupwindow())
    menuOptions.add_command(label="History", command= lambda:displaypreviousdata())
    menuOptions.add_command(label="End of Day", command= lambda:savefile())
    menuOptions.add_separator()
    menuOptions.add_command(label="Home", command= lambda:switchwindow())
    menuOptions.add_command(label="Exit", command= lambda:sys.exit())
    menuBar.add_cascade(label="More options",font=("Times new roman", 10, 'bold'), menu= menuOptions)
    root.config(menu=menuBar)


    frametop = Frame(root,height= 250, width= 850,  padx=20, pady=20, bg='aquamarine')
    frametop.grid( row=1, column= 0)

    framebottom = Frame(root,height= 250, width= 850, padx=30, pady=30,bg='aquamarine')
    framebottom.grid( row=2, column= 0)

    # Label Creation
    namelabel: Label = tk.Label(frametop, text="Product Name", font=('Times new roman', 18, 'bold'), bg='aquamarine')
    qrcodelabel: Label = tk.Label(frametop, text="QR Code", font=('Times new roman', 18, 'bold'), bg='aquamarine' )

    #Labels grid
    namelabel.grid(column=0, row=0, padx=10,pady=20)
    qrcodelabel.grid(column=0, row=1, padx=10, pady=20)

    # StringVar() to store the values from User input.
    name = StringVar()
    number = StringVar()
    qrcode = StringVar()
    status = StringVar()


    # Combobox creation
    nameComboBox = Combobox(frametop, state='readonly', textvariable=name, width=35,font=('Times new roman', 12, 'bold'), values=dic_names,
                            postcommand=lambda: selnumber())  # Put the values
    # numComboBox = Combobox(root, text-variable=number, width=35, values=dic_numbers)  # Put the values
    qrEntryBox = tk.Entry(frametop, textvariable=qrcode, width=38, font=('Times new roman', 12, 'bold'))

    #Entry grid
    nameComboBox.grid(column=2, row=0)
    qrEntryBox.grid(column=2, row=1)

    # Tree view


    style.configure("mystyle.Treeview", highlightthickness=50, bd=0, font=('Calibri', 12,'bold')) # Modify the font of the body
    style.configure("mystyle.Treeview.Heading", font=('Times new roman', 15,'bold')) # Modify the font of the headings
    style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) # Remove the borders
    
    displaydata = Treeview(framebottom, columns=['Product Name', 'Product Number', 'QR Code', 'Status'], show='headings',
                           height=15, selectmode='extended', takefocus="Any", style= 'mystyle.Treeview')
    
    
    displaydata.grid(column=0, row=1, rowspan= 4)
    displaydata.column("# 1", anchor=CENTER)
    displaydata.heading("# 1", text="Product Name")
    displaydata.column("# 2", anchor=CENTER)
    displaydata.heading("# 2", text="Product Number")
    displaydata.column("# 3", anchor=CENTER)
    displaydata.heading("# 3", text="QR Code")
    displaydata.column("# 4", anchor=CENTER)
    displaydata.heading("# 4", text="Status/Date")

    def cleartree():
         for item in displaydata.get_children():
          displaydata.delete(item)

         counterEntry.delete(0, END)

    cleartreebtn = Button(framebottom, text="Clear Data", relief= 'raised', font=('Times new roman', 15,'bold'), command= lambda : cleartree())
    cleartreebtn.grid(column= 0, row= 8, pady=20)

   


    # scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=displaydata.yview)
    # displaydata.configure(yscroll=scrollbar.set)
    # scrollbar.grid(column=30, row=270)

    def selnumber():
        if name.get() == "Black Plug Gateway, Ext Ant, CAN, Type A":
            number.set("GBP-2002-CAN-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type C":
            number.set("GBP-2002-EUC-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, EU, Type G":
            number.set("GBP-2002-EUG-EXA")
        if name.get() == "Black Plug Gateway, Ext Ant, US, ATT":
            number.set("GBP-2002-0A2-ATT")
        if name.get() == "Black Plug Gateway, Ext Ant, US, TMO":
            number.set("GBP-2002-0A2-TMO")
        if name.get() == "Black Plug Gateway, Ext Ant, US, VZN":
            number.set("GBP-2002-0A2-VZN")
        if name.get() == "Green Wallplug US":
            number.set("PGW-2003-0A1")
        if name.get() == "Green Wallplug, CAN, Type B":
            number.set("PGW-2003-CAN")
        if name.get() == "Green Wallplug, EU, Type E":
            number.set("PGW-2003-EUE")
        if name.get() == "Green Wallplug, EU, Type G":
            number.set("PGW-2003-EUG")
        if name.get() == "Black Plug Gateway, Ext Ant, ASIA, Type A":
            number.set("GBP-2002-UA96-X")
        if name.get() == "Black Plug Gateway, Ext Ant, IND, Type C":
            number.set("GBP-2002-EC80-X")
        if name.get() == "Black Plug Gateway, Ext Ant, HK, Type G":
            number.set("GBP-2002-EG82-X")
        if name.get() == "Black Plug Gateway, Ext Ant, NZ, Type I":
            number.set("GBP-2002-UI82-X")
        if name.get() == "Black Plug Gateway, Ext Ant, AUS, Type I":
            number.set("GBP-2002-UI96-X")
        if name.get() == "Black Plug Gateway, Ext Ant, SG, Type C":
            number.set("GBP-2002-UC96-X")
        if name.get() == "Black Plug Gateway, Ext Ant, KOR, Type C":
            number.set("GBP-2002-UC94-X")
        if name.get() == "Black Plug Gateway, Ext Ant, MEX, Type C":   
            number.set("GBP-2002-UC93-X")
        if name.get() == "Green Wallplug, ASIA, Type B":
            number.set("PGW-2003-UA96-I")
        if name.get() == "Green Wallplug, IND, Type E":
            number.set("PGW-2003-EE80-I")
        if name.get() == "Green Wallplug, HK, Type G":
            number.set("PGW-2003-EG82-I")
        if name.get() == "Green Wallplug NZ, Type I":
            number.set("PGW-2003-UI82-I")
        if name.get() == "Green Wallplug, AUS, Type I":
            number.set("PGW-2003-UI96-I")
        if name.get() == "Green Wallplug, SG, Type E":
            number.set("PGW-2003-UE96-I")
        if name.get() == "Green Wallplug, KOR, Type E":
            number.set("PGW-2003-UE94-I")
        if name.get() == "Green Wallplug, MEX, Type E":
            number.set("PGW-2003-UE93-I")

    


    def cleartext():
        qrEntryBox.delete(0, END)
        

    def datacheck():
        if qrcode.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No QR code entered')
            root.destroy()
            printing()
        elif name.get() == '':
            messagebox.showerror("Error! Missing Data.", 'No Product Name entered')
            root.destroy()
            printing()

    def save_to_excel_printlog():
        tstamp = str(datetime.now().strftime("%Y-%m-%d"))
        f = openpyxl.load_workbook('LabelPrintLog.xlsx')
        sheet = f.active
        maxrow = sheet.max_row

        data = ({"Product Name": [name.get()], "Product Number": [number.get()], "QR Code": [qrcode.get()],
                 "TimeStamp": [tstamp], "Status": [status.get()]})

        dataframe = pd.DataFrame(data)

        with pd.ExcelWriter('LabelPrintLog.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            dataframe.to_excel(writer, startrow=maxrow, startcol=0, header=None, index=False)

    

    def insertdata():

        datacheck()
        duplicatesprintLog()
        status.set("Printed")
        selnumber()
        displaydata.insert('', END,
                        values=[name.get(), number.get(), qrcode.get(), status.get()])
        printlabel()
        save_to_excel_printlog()  # Just for saving into a file SavedOnly xlsx.
        cleartext()
        counterVar.set(counter())
        
    root.bind('<Return>', lambda event: insertdata())




    def popupwindow():
        pilotname = simpledialog.askstring("Enter Name of the Product.", "Enter Name of the Product.", parent=root)
        name.set(pilotname)
        pilotnum = simpledialog.askstring("Enter number of the Product.", "Enter Number of the Product.", parent=root)
        number.set(pilotnum)

    # manualbtn = Button(root, text="Manual Entry", command=lambda: popupwindow(), font=('Times new roman', 15),
    #                    relief='ridge')
    #manualbtn.grid(column=3, row=0)

    def duplicatesprintLog():
        currentQrCode = qrcode.get()
        # currentName = name.get()
        # currentNumber = productNumber.get()
        wb = openpyxl.load_workbook('LabelPrintLog.xlsx', read_only=True)
        cs = wb.active

        for row in cs.rows:
            for cell in row:
                if currentQrCode == cell.value:
                    messagebox.showerror("Error", "Duplicate Entry!")
                    root.destroy()
                    printing()
                    

        wb.close()

    def printlabel():
        
        qrcodeType = qrcode.get()
        myPrinter = 'DYMO LabelWriter 450 Turbo'
        
        if qrcodeType[:2] == "14":

            url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + qrcode.get()
            r = requests.get(url)
            rawjson = json.loads(r.text)
            data = rawjson['data']
            imei = data[0]['IMEInumber']  

            qr_value = qrcode.get()
            qr_text = qrcode.get()
            barcode_value = number.get()
            text_value = name.get()
            qr_path = pathlib.Path('./US Green Wallplug Template with IMEI.label')

            printer_con = Dispatch('Dymo.DymoAddIn')
            printer_con.SelectPrinter(myPrinter)
            printer_con.Open(qr_path)
            printer_label = Dispatch('Dymo.DymoLabels')
            printer_label.SetField("BARCODE_1", qr_value)
            printer_label.SetField("TEXT_1", qr_text)
            printer_label.SetField("BARCODE", barcode_value)
            printer_label.SetField("TEXT", text_value)
            printer_label.SetField("BARCODE_2", imei)
            printer_con.setGraphicsAndBarcodePrintMode(OFF)

            printer_con.StartPrintJob()
            printer_con.Print(1, False)
            printer_con.EndPrintJob()

        else:


            # url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + qrcode.get()
            # r = requests.get(url)
            # rawjson = json.loads(r.text)
            # data = rawjson['data']
            # imei = data[0]['IMEInumber']  

            qr_value = qrcode.get()
            qr_text = qrcode.get()
            barcode_value = number.get()
            text_value = name.get()
            qr_path = pathlib.Path('./US Green Wallplug Template.label')

            printer_con = Dispatch('Dymo.DymoAddIn')
            printer_con.SelectPrinter(myPrinter)
            printer_con.Open(qr_path)
            printer_label = Dispatch('Dymo.DymoLabels')
            printer_label.SetField("BARCODE_1", qr_value)
            printer_label.SetField("TEXT_1", qr_text)
            printer_label.SetField("BARCODE", barcode_value)
            printer_label.SetField("TEXT", text_value)
            # printer_label.SetField("BARCODE_2", imei)
            printer_con.setGraphicsAndBarcodePrintMode(OFF)

            printer_con.StartPrintJob()
            printer_con.Print(1, False)
            printer_con.EndPrintJob()

    def switchwindow():
        root.destroy()
        import main
        main.main()
    
    # photo = PhotoImage(file='Home_icon_black.png')
    # switchbtn = Button(root, image=photo, height=40, width=40, text="Home", command=lambda: switchwindow() )
    # # switchbtn.grid(column=0, row=0)



    def reprintItems():
        curItem = displaydata.focus()
        name.set(displaydata.item(curItem)['values'][0])
        number.set(displaydata.item(curItem)['values'][1])
        qrcode.set(displaydata.item(curItem)['values'][2])
        datacheck()
        status.set("Reprinted")
        selnumber()
        displaydata.insert('', END,
                           values=[name.get(), number.get(), qrcode.get(), status.get()])
        printlabel()
       # save_to_excel_printlog()  # Just for saving into a file SavedOnly xlsx.
        cleartext()

    reprintbtn = Button(frametop, text="Reprint",relief= 'raised', font=('Times new roman', 13,'bold'), height=1, command=lambda: reprintItems())
    Label(frametop, text="\t", bg="aquamarine").grid(column=3, row=0)
    # Label(frametop, text="\t").grid(column=3, row=1)
    reprintbtn.grid(column=4, row=0, sticky="s")


    def displaypreviousdata():

        file = filedialog.askopenfilename(title="Open a File", filetype=(("xlsx files", ".*xlsx"), ("All Files", "*.")))

        df = pd.read_excel(file)

        for item in displaydata.get_children():
            displaydata.delete(item)

        for cell in range(len(df)):
            displaydata.insert("", END, values=list(df.loc[cell]))

    # historybtn = Button(root, text="History", command=lambda: displaypreviousdata(), font=('Times new roman', 10),
    #                   relief='ridge')
    # historybtn.grid(column=0 , row=1)

    def savefile():

        msgbox = messagebox.askokcancel("End of day.", "Are you sure you want to perform End of day?")

        if msgbox == True:

            tstamp = str(datetime.now().strftime("%Y-%m-%d"))
            filename = f"Data of {tstamp}.xlsx"

            if exists(f"LOG\\{filename}") != True:

                with open(f"LOG\\{filename}", 'a') as f:


                    # destination = filedialog.askopenfilename(initialdir="LOG\\")
                    destinationfile = f"LOG\\{filename}"

                    source = "LabelPrintLog.xlsx"
                    sourcefile = os.path.abspath(source)

                    shutil.copy(sourcefile, destinationfile)
                    print("File Copied!")
            else:
                messagebox.showerror("File Error","File Already Exists. Creating a Copy")
                with open(f"LOG\\Copy of {filename}", 'a') as f:


                    destination = filedialog.askopenfilename(initialdir="LOG\\")
                    destinationfile = os.path.abspath(destination)

                    source = "LabelPrintLog.xlsx"
                    sourcefile = os.path.abspath(source)

                    shutil.copy(sourcefile, destinationfile)
                    print("File Copied!")

        # source = openpyxl.load_workbook("LabelPrintLog.xlsx")
        # sheet = source.active


    # eodBtn = Button(root, text="End Of Day", command=lambda: savefile(), font=('Times new roman', 10),
    #                     relief='ridge')
    #  savebtn.grid(column=0, row=2)
    
    # date_Today = datetime.today()

    # def createfile():
        
    #     current_directory = filedialog.askdirectory()
    #     today_date = date_Today.strftime("%row%m%d")
    #     filename = "Data of "+today_date + ".xlsx"
    #     if os.path.exists(filename) == True:
    #         messagebox.showerror("Warning", "This file name already exist. Please enter a different name.")
    #         filename = simpledialog.askstring(title = "FileName", prompt=f"Current file name is {filename}",) + ".xlsx"
    #         newfile = openpyxl.Workbook()
    #         newfile.create_sheet("Sheet_one")
    #         newfile.save(filename)
    #     else:
    #         newfile = openpyxl.Workbook()
    #         newfile.create_sheet("Sheet_one")
    #         newfile.save(filename)
        
    def counter():
        df = pd.read_excel('LabelPrintLog.xlsx')
        countTotal = str(len(df.index))
        return countTotal
        
        
    counterVar = StringVar()
    totalprintlabel = Label(framebottom, text=f"Total Printed").grid(column=0, row=6, pady=10)
    counterEntry = Entry(framebottom, text=counterVar, width=3, font=("Times new Roman", 14, 'bold'), state='readonly')
    counterEntry.grid(column=0, row=7)
    
    
    root.mainloop()

